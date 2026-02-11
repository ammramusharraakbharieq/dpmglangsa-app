"""Check Excel headers for perangkat desa"""
from openpyxl import load_workbook
from pathlib import Path

BASE_DIR = Path(__file__).parent
file_path = BASE_DIR / 'data_(kepala desa & perangkat desa).xlsx'

wb = load_workbook(file_path)
ws = wb.active

print("Header rows (1-5):")
for row in range(1, 6):
    print(f"Row {row}:")
    for col in range(1, 17):
        val = ws.cell(row=row, column=col).value
        if val:
            print(f"  Col {col}: {val}")

print("\nData row 6 (first data row):")
for col in range(1, 17):
    val = ws.cell(row=6, column=col).value
    print(f"  Col {col}: {val}")

wb.close()
