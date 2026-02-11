"""Verify all kecamatan headers are now correct"""
import openpyxl
from pathlib import Path

file_path = Path(__file__).parent / "data_(kepala desa & perangkat desa).xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Find header rows by merged K:K+1 ranges
header_rows = []
for mc in ws.merged_cells.ranges:
    if mc.min_col == 11 and mc.max_col == 11 and (mc.max_row - mc.min_row) == 1:
        header_rows.append(mc.min_row)
header_rows.sort()

expected_col11_16 = {
    11: 'NO', 12: 'NAMA LENGKAP',
    13: 'NOMOR INDUK KEPENDUDUKAN ( NIK)',
    14: 'JENIS KELAMIN', 15: 'JABATAN', 16: 'NOMOR HP'
}

all_ok = True
for hr in header_rows:
    # Find kecamatan name
    kec_name = None
    for data_row in range(hr + 3, min(hr + 10, ws.max_row + 1)):
        val = ws.cell(row=data_row, column=6).value
        if val:
            kec_name = str(val).strip()
            break
    
    print(f"\n{'='*60}")
    print(f"Header at row {hr} - Kecamatan: {kec_name}")
    print(f"{'='*60}")
    
    # Check columns 11-16
    for col, expected in expected_col11_16.items():
        actual = ws.cell(row=hr, column=col).value
        status = "OK" if actual == expected else f"MISMATCH (got: '{actual}')"
        if actual != expected:
            all_ok = False
        print(f"  Col {col}: {status} -> '{actual}'")
    
    # Check sub-header row
    sub_row = hr + 1
    print(f"\n  Sub-header row {sub_row}:")
    for col in range(1, 9):
        val = ws.cell(row=sub_row, column=col).value
        print(f"    Col {col}: '{val}'")
    
    # Check column number row
    num_row = hr + 2
    print(f"\n  Column numbers row {num_row}:")
    vals = [ws.cell(row=num_row, column=c).value for c in range(1, 17)]
    print(f"    {vals}")

print(f"\n{'='*60}")
print(f"Overall: {'ALL OK!' if all_ok else 'SOME ISSUES FOUND'}")
print(f"{'='*60}")

wb.close()
