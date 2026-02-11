"""Diagnose which cells are MergedCell objects near header boundaries"""
import openpyxl
from pathlib import Path
from openpyxl.cell.cell import MergedCell

file_path = Path(__file__).parent / "data_(kepala desa & perangkat desa).xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Find all header rows  
header_rows = []
for mc in ws.merged_cells.ranges:
    if mc.min_col == 11 and mc.max_col == 11 and (mc.max_row - mc.min_row) == 1:
        header_rows.append(mc.min_row)
header_rows.sort()

print("Header rows:", header_rows)

# For each header, check 3 rows above and the header rows themselves for MergedCell
for hr in header_rows:
    print(f"\n--- Header at row {hr} ---")
    for r in range(max(1, hr - 3), hr + 4):
        merged_cols = []
        for col in range(1, 17):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell, MergedCell):
                merged_cols.append(col)
        if merged_cols:
            print(f"  Row {r}: MergedCell at columns {merged_cols}")
        else:
            vals = {}
            for col in range(1, 17):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    vals[col] = v
            print(f"  Row {r}: Regular cells - {vals if vals else 'EMPTY'}")

# Also check which merged ranges exist
print("\n\n--- All merged cell ranges ---")
for mc in sorted(ws.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col)):
    val = ws.cell(row=mc.min_row, column=mc.min_col).value
    print(f"  {mc} -> '{val}'")

# Check what happens around the last desa in each kecamatan section
print("\n\n--- Last rows before each header ---")
for hr in header_rows:
    if hr <= 3:
        continue
    print(f"\n  Before header at row {hr}:")
    for r in range(max(1, hr-5), hr):
        vals = {}
        is_mc = []
        for col in range(1, 17):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell, MergedCell):
                is_mc.append(col)
            elif cell.value is not None:
                vals[col] = cell.value
        status = f"MergedCell@{is_mc}" if is_mc else ""
        print(f"    Row {r}: {status} {vals if vals else 'EMPTY'}")

wb.close()
