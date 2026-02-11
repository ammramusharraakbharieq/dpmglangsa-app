"""Analyze Excel header structure and merged cells to find broken headers"""
import openpyxl
from pathlib import Path

file_path = Path(r"d:\sistem-dpmg-langsa") / "data_(kepala desa & perangkat desa).xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("=" * 80)
print("MERGED CELLS ANALYSIS")
print("=" * 80)

# Show all merged cell ranges sorted by row
merged = sorted(ws.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col))
print(f"\nTotal merged cells: {len(merged)}")
for mc in merged:
    # Get the value in the top-left cell
    val = ws.cell(row=mc.min_row, column=mc.min_col).value
    print(f"  {mc} -> value='{val}'")

print("\n" + "=" * 80)
print("ALL ROWS WITH DATA - Full dump")
print("=" * 80)

# Dump all rows that have any content
for row in range(1, ws.max_row + 1):
    vals = {}
    for col in range(1, 17):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            vals[col] = v
    if vals:
        print(f"Row {row}: {vals}")

print("\n" + "=" * 80)
print("KECAMATAN SECTIONS - Header rows analysis")
print("=" * 80)

# Find where each kecamatan starts by looking at column 6
kecamatan_names = ['LANGSA TIMUR', 'LANGSA KOTA', 'LANGSA BARAT', 'LANGSA BARO', 'LANGSA LAMA']
kecamatan_rows = {}

for row in range(1, ws.max_row + 1):
    for col in range(1, 17):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            val_upper = str(val).strip().upper()
            for kec in kecamatan_names:
                if kec == val_upper:
                    if kec not in kecamatan_rows:
                        kecamatan_rows[kec] = row

print(f"\nKecamatan first occurrence rows:")
for kec, r in sorted(kecamatan_rows.items(), key=lambda x: x[1]):
    print(f"  {kec}: row {r}")
    # Show 4 rows above this for header context
    for hr in range(max(1, r-4), r+2):
        vals = {}
        for col in range(1, 17):
            v = ws.cell(row=hr, column=col).value
            if v is not None:
                vals[col] = v
        print(f"    Row {hr}: {vals}")

wb.close()
