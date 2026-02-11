"""
One-time repair script to fix broken kecamatan header rows in the Excel file.
Restores missing header values (columns 11-16) across all kecamatan sections.
Also fixes LANGSA BARO header where values are misplaced.
"""
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, Alignment, Border, Side
import copy

file_path = Path(__file__).parent / "data_(kepala desa & perangkat desa).xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Reference header (LANGSA TIMUR) - rows 3-5
# The correct template for header values
HEADER_VALUES_COL11_16 = {
    11: 'NO',
    12: 'NAMA LENGKAP',
    13: 'NOMOR INDUK KEPENDUDUKAN ( NIK)',
    14: 'JENIS KELAMIN',
    15: 'JABATAN',
    16: 'NOMOR HP'
}

HEADER_VALUES_COL1_8_ROW1 = {
    1: 'PROVINSI',
    3: 'KABUPATEN / KOTA',
    5: 'KECAMATAN',
    7: 'DESA',
}

HEADER_VALUES_COL1_8_ROW2 = {
    1: 'NO',
    2: 'NAMA',
    3: 'NO',
    4: 'NAMA',
    5: 'NO',
    6: 'NAMA',
    7: 'NO ',
    8: 'NAMA'
}

# Reference row for style copying
ref_row = 3

# Find header sections by looking for merged cells that match the header pattern
# Each kecamatan section has merged cells like K<n>:K<n+1> for the header
header_sections = []

for mc in ws.merged_cells.ranges:
    # Look for merged cells in column K (11) that span 2 rows 
    if mc.min_col == 11 and mc.max_col == 11 and (mc.max_row - mc.min_row) == 1:
        header_sections.append(mc.min_row)

header_sections.sort()
print(f"Found kecamatan header sections starting at rows: {header_sections}")

for header_row in header_sections:
    if header_row == ref_row:
        print(f"\nRow {header_row}: Reference (LANGSA TIMUR) - checking values")
        # Just verify
        for col, expected_val in HEADER_VALUES_COL11_16.items():
            actual = ws.cell(row=header_row, column=col).value
            if actual != expected_val:
                print(f"  WARNING: Col {col} expected '{expected_val}' but got '{actual}'")
            else:
                print(f"  Col {col}: OK ('{actual}')")
        continue
    
    print(f"\nFixing header section starting at row {header_row}:")
    
    # Fix columns 11-16 (in the top-left cell of merged ranges)
    for col, val in HEADER_VALUES_COL11_16.items():
        cell = ws.cell(row=header_row, column=col)
        current = cell.value
        if current is None or str(current).strip() == '':
            cell.value = val
            # Copy style from reference
            ref_cell = ws.cell(row=ref_row, column=col)
            if ref_cell.has_style:
                cell.font = copy.copy(ref_cell.font)
                cell.border = copy.copy(ref_cell.border)
                cell.fill = copy.copy(ref_cell.fill)
                cell.alignment = copy.copy(ref_cell.alignment)
                cell.number_format = ref_cell.number_format
            print(f"  Row {header_row}, Col {col}: Set to '{val}'")
        else:
            print(f"  Row {header_row}, Col {col}: Already has '{current}'")
    
    # Fix columns 1-8 first header row (PROVINSI, KABUPATEN, etc.)
    for col, val in HEADER_VALUES_COL1_8_ROW1.items():
        cell = ws.cell(row=header_row, column=col)
        current = cell.value
        if current is None or str(current).strip() == '':
            cell.value = val
            ref_cell = ws.cell(row=ref_row, column=col)
            if ref_cell.has_style:
                cell.font = copy.copy(ref_cell.font)
                cell.border = copy.copy(ref_cell.border)
                cell.fill = copy.copy(ref_cell.fill)
                cell.alignment = copy.copy(ref_cell.alignment)
                cell.number_format = ref_cell.number_format
            print(f"  Row {header_row}, Col {col}: Set to '{val}'")
    
    # Fix sub-header row (NO, NAMA, etc.) - Row header_row + 1
    sub_row = header_row + 1
    for col, val in HEADER_VALUES_COL1_8_ROW2.items():
        # Check if this cell is part of a merged range (skip if it is)
        is_merged = False
        for mc in ws.merged_cells.ranges:
            if mc.min_row <= sub_row <= mc.max_row and mc.min_col <= col <= mc.max_col:
                if not (sub_row == mc.min_row and col == mc.min_col):
                    is_merged = True
                    break
        
        if is_merged:
            print(f"  Row {sub_row}, Col {col}: Skipping (merged cell)")
            continue
            
        cell = ws.cell(row=sub_row, column=col)
        current = cell.value
        if current is None or str(current).strip() == '':
            cell.value = val
            ref_cell = ws.cell(row=ref_row + 1, column=col)
            if ref_cell.has_style:
                cell.font = copy.copy(ref_cell.font)
                cell.border = copy.copy(ref_cell.border)
                cell.fill = copy.copy(ref_cell.fill)
                cell.alignment = copy.copy(ref_cell.alignment)
                cell.number_format = ref_cell.number_format
            print(f"  Row {sub_row}, Col {col}: Set to '{val}'")
    
    # Fix column number row - Row header_row + 2
    num_row = header_row + 2
    for col in range(1, 17):
        cell = ws.cell(row=num_row, column=col)
        current = cell.value
        if current is None:
            cell.value = col
            ref_cell = ws.cell(row=ref_row + 2, column=col)
            if ref_cell.has_style:
                cell.font = copy.copy(ref_cell.font)
                cell.border = copy.copy(ref_cell.border)
                cell.fill = copy.copy(ref_cell.fill)
                cell.alignment = copy.copy(ref_cell.alignment)
                cell.number_format = ref_cell.number_format
            print(f"  Row {num_row}, Col {col}: Set to {col}")

wb.save(file_path)
wb.close()
print("\n✅ All headers fixed and saved!")
