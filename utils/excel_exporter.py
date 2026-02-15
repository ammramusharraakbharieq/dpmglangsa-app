
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from pathlib import Path
import shutil

class ExcelExporter:
    def __init__(self, base_path):
        self.base_path = Path(base_path)

    def _load_template(self, filename):
        """Load the workbook from file, effectively using it as a template"""
        file_path = self.base_path / filename
        if not file_path.exists():
            raise FileNotFoundError(f"Template file not found: {filename}")
        
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        return wb

    def _clear_data(self, ws, start_row, end_col):
        """Clear data from start_row downwards, preserving row structure/style"""
        from openpyxl.cell import MergedCell
        
        # KEY FIX 2: Explicitly find all merged ranges in data area and unmerge them
        # We must collect them first to avoid iterator modification issues
        ranges_to_unmerge = []
        for merged_range in list(ws.merged_cells.ranges): # Use list() to copy
            if merged_range.min_row >= start_row:
                ranges_to_unmerge.append(merged_range)
        
        for merged_range in ranges_to_unmerge:
            try:
                ws.unmerge_cells(str(merged_range))
            except KeyError:
                pass # Already unmerged or not found

        max_row = ws.max_row
        if max_row >= start_row:
            for row in range(start_row, max_row + 1):
                for col in range(1, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    # Double check: if it's still a merged cell (e.g. from above header), skip it
                    if isinstance(cell, MergedCell):
                        continue # Don't touch header-merged cells!
                    cell.value = None

    def export_camat_mukim_geuchik(self, df):
        filename = "data_(camat,mukim,dan geuchik).xlsx"
        wb = self._load_template(filename)
        ws = wb.active
        start_row = 2 # Start clearing from Row 2

        self._clear_data(ws, start_row, 7)

        # Columns mapping:
        # A: NO, B: KECAMATAN, C: NAMA CAMAT, D: KEMUKIMAN, E: NAMA MUKIM, F: GAMPONG, G: NAMA GEUCHIK
        
        # Ensure DF has these columns
        # DF from load_camat_mukim_geuchik has:
        # NO, KECAMATAN, NAMA_CAMAT, KEMUKIMAN, NAMA_MUKIM, GAMPONG, NAMA_GEUCHIK
        
        for index, row in df.iterrows():
            current_row = start_row + index
            ws.cell(row=current_row, column=1, value=index + 1) # NO
            ws.cell(row=current_row, column=2, value=row.get('KECAMATAN'))
            # Corrected keys based on data_loader (from Excel Header)
            ws.cell(row=current_row, column=3, value=row.get('NAMA CAMAT'))
            ws.cell(row=current_row, column=4, value=row.get('KEMUKIMAN'))
            ws.cell(row=current_row, column=5, value=row.get('NAMA MUKIM'))
            ws.cell(row=current_row, column=6, value=row.get('GAMPONG'))
            ws.cell(row=current_row, column=7, value=row.get('NAMA GEUCHIK'))

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_geuchik_detail(self, df):
        filename = "data_(geuchik kota langsa).xlsx"
        wb = self._load_template(filename)
        ws = wb.active
        start_row = 4

        self._clear_data(ws, start_row, 18)

        # Mapping based on analysis and data_loader
        # A(1): PROVINSI (NO in df?), 
        # B(2): PROVINSI NAME? -> Header says PROVINSI merged? 
        # Wait, let's look at row 4 of analysis:
        # Col 0 ('11'): NO_PROV? 
        # Col 1 ('ACEH'): PROVINSI
        # Col 2 ('11.74'): NO_KAB
        # Col 3 ('LANGSA'): KABUPATEN
        # Col 4 ('11.74.01'): NO_KEC
        # Col 5 ('LANGSA TIMUR'): KECAMATAN
        # Col 6 ('11.74.01.2007'): NO_DESA
        # Col 7 ('BUKET MEDANG ARA'): DESA
        # Col 8 ('MUHAMMAD JAMIL...'): NAMA_LENGKAP
        # Col 9: TGL_LAHIR
        # Col 10: BLN_LAHIR
        # Col 11: THN_LAHIR
        # Col 12: JENIS_KELAMIN
        # Col 13: PENDIDIKAN
        # Col 14: SK_NOMOR
        # Col 15: SK_TANGGAL
        # Col 16: JABATAN
        # Col 17: NO_HP
        
        for index, row in df.iterrows():
            r = start_row + index
            # data_loader.load_geuchik_detail columns:
            # NO_PROV, PROVINSI, NO_KAB, KABUPATEN, NO_KEC, KECAMATAN, NO_DESA, DESA, 
            # NAMA_LENGKAP, TGL_LAHIR, BLN_LAHIR, THN_LAHIR, JENIS_KELAMIN, PENDIDIKAN,
            # SK_NOMOR, SK_TANGGAL, JABATAN, NO_HP
            
            ws.cell(row=r, column=1, value=row.get('NO_PROV'))
            ws.cell(row=r, column=2, value=row.get('PROVINSI'))
            ws.cell(row=r, column=3, value=row.get('NO_KAB'))
            ws.cell(row=r, column=4, value=row.get('KABUPATEN'))
            ws.cell(row=r, column=5, value=row.get('NO_KEC'))
            ws.cell(row=r, column=6, value=row.get('KECAMATAN'))
            ws.cell(row=r, column=7, value=row.get('NO_DESA'))
            ws.cell(row=r, column=8, value=row.get('DESA'))
            ws.cell(row=r, column=9, value=row.get('NAMA_LENGKAP'))
            ws.cell(row=r, column=10, value=row.get('TGL_LAHIR'))
            ws.cell(row=r, column=11, value=row.get('BLN_LAHIR'))
            ws.cell(row=r, column=12, value=row.get('THN_LAHIR'))
            ws.cell(row=r, column=13, value=row.get('JENIS_KELAMIN'))
            ws.cell(row=r, column=14, value=row.get('PENDIDIKAN'))
            ws.cell(row=r, column=15, value=row.get('SK_NOMOR'))
            ws.cell(row=r, column=16, value=row.get('SK_TANGGAL'))
            ws.cell(row=r, column=17, value=row.get('JABATAN'))
            ws.cell(row=r, column=18, value=row.get('NO_HP'))

        output = BytesIO()
        wb.save(output)
        return output.getvalue()


    def export_perangkat_desa(self, df):
        # PROGRAMMATIC GENERATION (Replacing Template)
        # To ensure clean layout matching "Perangkat_Desa" Google Sheet structure
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Perangkat_Desa"
        
        # --- STYLES ---
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        bold_font = Font(bold=True, name='Arial', size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        # --- HEADERS ---
        # Based on Analysis of Google Sheet structure
        
        # Row 1-2: Main Titles (Merged)
        ws.merge_cells('A1:P1')
        ws['A1'] = "DATA KEPALA DESA DAN PERANGKAT DESA"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align
        
        ws.merge_cells('A2:P2')
        ws['A2'] = "PEMERINTAH KOTA LANGSA" 
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = center_align

        # Row 3: Column Headers
        headers = [
            "NO PROV", "PROVINSI", "NO KAB", "KABUPATEN / KOTA", 
            "NO KEC", "KECAMATAN", "NO DESA", "DESA", 
            "KATEGORI", "JENIS", "NO URUT", "NAMA LENGKAP", 
            "NIK", "JENIS KELAMIN", "JABATAN", "NOMOR HP"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
            # Add light gray background for headers
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Row 4: Column Numbers (1-16)
        for col_num in range(1, 17):
            cell = ws.cell(row=4, column=col_num)
            cell.value = col_num
            cell.font = Font(italic=True, size=9)
            cell.alignment = center_align
            cell.border = thin_border

        # --- DATA PREPARATION ---
        # Filter "Ghost Rows" where Name and Jabatan are empty
        if 'NAMA_LENGKAP' in df.columns and 'JABATAN' in df.columns:
             name_str = df['NAMA_LENGKAP'].fillna('').astype(str).str.strip()
             jabatan_str = df['JABATAN'].fillna('').astype(str).str.strip()
             df = df[(name_str != '') | (jabatan_str != '')]

        # Convert NO_URUT to numeric for sorting
        df['NO_URUT_NUM'] = pd.to_numeric(df['NO_URUT'], errors='coerce')
        
        # Sort by Hierarchy: KEC -> DESA -> NO_URUT
        sort_cols = []
        if 'NO_KEC' in df.columns: sort_cols.append('NO_KEC')
        if 'NO_DESA' in df.columns: sort_cols.append('NO_DESA')
        if 'DESA' in df.columns: sort_cols.append('DESA') 
        sort_cols.append('NO_URUT_NUM')
        
        df = df.sort_values(by=sort_cols)
        
        # --- WRITE DATA ---
        start_row = 5
        current_row = start_row
        
        # Grouping Logic
        last_desa = None
        village_ranges = []
        current_village_start = start_row
        
        for index, row in df.iterrows():
            desa = row.get('DESA')
            
            # Detect Village Change
            if desa != last_desa:
                if last_desa is not None:
                     village_ranges.append({
                         'start': current_village_start,
                         'end': current_row - 1
                     })
                current_village_start = current_row
                last_desa = desa
                
                # Write Village Info (Cols 1-8) only on first row of group
                ws.cell(row=current_row, column=1, value=row.get('NO_PROV'))
                ws.cell(row=current_row, column=2, value=row.get('PROVINSI'))
                ws.cell(row=current_row, column=3, value=row.get('NO_KAB'))
                ws.cell(row=current_row, column=4, value=row.get('KABUPATEN'))
                ws.cell(row=current_row, column=5, value=row.get('NO_KEC'))
                ws.cell(row=current_row, column=6, value=row.get('KECAMATAN'))
                ws.cell(row=current_row, column=7, value=row.get('NO_DESA'))
                ws.cell(row=current_row, column=8, value=desa)
            
            # Helper for Kategori (Col 9) and Jenis (Col 10)
            jabatan = str(row.get('JABATAN', '')).upper()
            kategori = 'B' # Default
            jenis = 'Prangkat'
            
            if any(x in jabatan for x in ['KEPALA DESA', 'PJ. KEPALA DESA']):
                kategori = 'A'
                jenis = 'Kades'
            
            # Write row data
            # If same village, Cols 1-8 are left empty (will be merged)
            
            ws.cell(row=current_row, column=9, value=kategori)
            ws.cell(row=current_row, column=10, value=jenis)
            ws.cell(row=current_row, column=11, value=row.get('NO_URUT'))
            ws.cell(row=current_row, column=12, value=row.get('NAMA_LENGKAP'))
            
            nik_val = row.get('NIK')
            ws.cell(row=current_row, column=13, value=str(nik_val) if pd.notna(nik_val) else '')
            
            ws.cell(row=current_row, column=14, value=row.get('JENIS_KELAMIN'))
            ws.cell(row=current_row, column=15, value=row.get('JABATAN'))
            
            hp_val = row.get('NO_HP')
            ws.cell(row=current_row, column=16, value=str(hp_val) if pd.notna(hp_val) else '')
            
            # Apply Borders and Alignment to this row
            for col in range(1, 17):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                # Center align most cols, Left align Name and Jabatan if desired?
                # For now match template: usually center or left. 
                # Let's Center basic info, Left align Name/Jabatan
                if col in [12, 15]: # Name, Jabatan
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

            current_row += 1

        # Append last group
        if last_desa is not None:
             village_ranges.append({
                 'start': current_village_start,
                 'end': current_row - 1
             })

        # --- POST-PROCESSING: MERGE CELLS ---
        for grp in village_ranges:
            start = grp['start']
            end = grp['end']
            
            # Merge Cols 1-8 for the group
            if end >= start: # Use >= so even single rows get processed (though merge start=end is no-op usually, alignment is key)
                for col in range(1, 9): # Cols A-H
                    if end > start:
                        ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)
                    
                    # Ensure alignment for the top-left cell of merge
                    cell = ws.cell(row=start, column=col)
                    cell.alignment = center_align
        
        # --- COLUMN WIDTHS (Auto-ish) ---
        column_widths = [
            10, 15, 10, 20, 12, 20, 15, 25, # 1-8
            8, 10, 8, 35, 20, 12, 30, 18  # 9-16
        ]
        for i, width in enumerate(column_widths, 1):
             ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_tuha_peuet(self, df):
        filename = "data_(tuha peuet gampong).xlsx"
        wb = self._load_template(filename)
        ws = wb.active
        start_row = 8
        self._clear_data(ws, start_row, 11)

        # Cols:
        # A(1): NO (Global Sequential 1,1,1..?)
        # Wait, Analysis Row 8: '1', 'LANGSA TIMUR', '1', 'SEUNEUBOK ANTARA', '1', 'BUKET MEDANG ARA', '1', 'Nurhasan'...
        # Row 9: '', '', '', '', '', '', '2', 'Bustami'...
        # So Cols A-F are Grouping Cols.
        # Col G(7): No Anggota (1,2,3...)
        # Col H(8): Nama
        # Col I(9): L (Checkmark)
        # Col J(10): P (Checkmark)
        # Col K(11): Ket
        
        current_row = start_row
        last_gampong = None
        global_no = 1
        
        # We need external counters for grouping
        # But looking at analysis:
        # Row 8 (New Gampong):
        # A: 1 (Global No?)
        # B: LANGSA TIMUR
        # C: 1 (No Mukim?)
        # D: SEUNEUBOK ANTARA (Mukim)
        # E: 1 (No Gampong?)
        # F: BUKET MEDANG ARA (Gampong)
        
        for index, row in df.iterrows():
            gampong = row.get('GAMPONG')
            is_new_gampong = (gampong != last_gampong)
            
            if is_new_gampong:
                # We assume simple global increment for A? 
                # Or is it per Kecamatan? Analysis shows '1' at start.
                # Let's just use an incrementing integer for now.
                ws.cell(row=current_row, column=1, value=global_no)
                ws.cell(row=current_row, column=2, value=row.get('KECAMATAN'))
                ws.cell(row=current_row, column=3, value=row.get('NO_KEMUKIMAN'))
                ws.cell(row=current_row, column=4, value=row.get('KEMUKIMAN'))
                ws.cell(row=current_row, column=5, value=row.get('NO_GAMPONG'))
                ws.cell(row=current_row, column=6, value=gampong)
                
                global_no += 1
                last_gampong = gampong
            
            # Data Columns (G-K)
            ws.cell(row=current_row, column=7, value=row.get('NO_ANGGOTA'))
            ws.cell(row=current_row, column=8, value=row.get('NAMA_ANGGOTA'))
            
            # Checkbox Logic
            # Check input from DF (LAKI_LAKI / PEREMPUAN cols)
            # If they contain '✓' or specific codes, preserve them.
            # If empty but we have JENIS_KELAMIN info? (Helper might need to add JK column to Load if missing)
            # Currently load_tuha_peuet returns LAKI_LAKI and PEREMPUAN columns directly from sheet.
            # So we just pass them through.
            
            ws.cell(row=current_row, column=9, value=row.get('LAKI_LAKI'))
            ws.cell(row=current_row, column=10, value=row.get('PEREMPUAN'))
            ws.cell(row=current_row, column=11, value=row.get('KETERANGAN'))
            
            current_row += 1

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
